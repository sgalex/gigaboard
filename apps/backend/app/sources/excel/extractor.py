"""Excel Source Extractor - multi-sheet with AI extraction code.

Поддержка .xlsx файлов с несколькими листами.
AI генерирует Python код для сложного извлечения и объединения данных.
"""
import time
from typing import Any

from app.sources.base import (
    BaseSource,
    ExtractionResult,
    ValidationResult,
    TableData,
)


class ExcelSource(BaseSource):
    """Excel file source handler."""
    
    source_type = "excel"
    display_name = "Excel"
    icon = "📗"
    description = "Excel файлы с несколькими листами"
    
    async def validate_config(self, config: dict[str, Any]) -> ValidationResult:
        """Validate Excel source config."""
        errors = []
        
        if not config.get("file_id") and not config.get("filename"):
            errors.append("Необходимо указать file_id или filename")
        
        if errors:
            return ValidationResult.failure(errors)
        return ValidationResult.success()
    
    async def extract(
        self,
        config: dict[str, Any],
        file_content: bytes | None = None,
        **kwargs
    ) -> ExtractionResult:
        """Extract data from Excel file.
        
        Два режима извлечения:
        
        1. **Простой** (analysis_mode='simple' или отсутствует):
           - sheets: list[str] — листы (пусто = все)
           - has_header: bool — первая строка — заголовки
           - max_rows: int | None — лимит строк
           - selected_columns: dict[sheet, list[col]] — столбцы
        
        2. **Умный** (analysis_mode='smart'):
           - detected_regions: list[dict] — регионы от ExcelTableDetector
             Каждый регион: sheet_name, start_row, start_col, end_row, end_col,
             header_row, table_name, selected_columns
        
        См. docs/SOURCE_NODE_CONCEPT_V2.md - "Excel Dialog".
        """
        start_time = time.time()
        
        if file_content is None:
            return ExtractionResult.failure("Содержимое файла не предоставлено")
        
        try:
            import io
            import pandas as pd
            
            excel_file = io.BytesIO(file_content)
            xlsx = pd.ExcelFile(excel_file)
            all_sheet_names = xlsx.sheet_names
            tables = []
            
            analysis_mode = config.get("analysis_mode", "simple")
            has_header = config.get("has_header", True)
            max_rows = config.get("max_rows")
            
            if analysis_mode == "smart" and config.get("detected_regions"):
                # ─── Smart mode: извлечение по регионам ─────────────────
                tables = self._extract_regions(
                    xlsx, config["detected_regions"], has_header, max_rows
                )
            else:
                # ─── Simple mode: листы целиком ─────────────────────────
                tables = self._extract_sheets(
                    xlsx, all_sheet_names, config, has_header, max_rows
                )
            
            extraction_time = int((time.time() - start_time) * 1000)
            total_rows = sum(len(t.rows) for t in tables)
            
            return ExtractionResult(
                success=True,
                text=f"Excel файл '{config.get('filename', 'data.xlsx')}': {len(tables)} таблиц, {total_rows} строк.",
                tables=tables,
                extraction_time_ms=extraction_time,
                metadata={
                    "sheet_names": all_sheet_names,
                    "selected_sheets": list({t.name.split(" :: ")[0] for t in tables}),
                    "table_count": len(tables),
                    "total_rows": total_rows,
                    "analysis_mode": analysis_mode,
                }
            )
            
        except ImportError:
            return ExtractionResult.failure("Необходимо установить pandas и openpyxl")
        except Exception as e:
            return ExtractionResult.failure(f"Ошибка чтения Excel: {str(e)}")
    
    # ─── Simple mode extraction ──────────────────────────────────────
    
    def _extract_sheets(
        self,
        xlsx,
        all_sheet_names: list[str],
        config: dict[str, Any],
        has_header: bool,
        max_rows: int | None,
    ) -> list[TableData]:
        """Extract entire sheets (simple / legacy mode)."""
        import pandas as pd
        
        selected_sheets = config.get("sheets", [])
        selected_columns_map = config.get("selected_columns", {})
        sheets_to_process = selected_sheets if selected_sheets else all_sheet_names
        tables: list[TableData] = []
        
        for sheet_name in sheets_to_process:
            if sheet_name not in all_sheet_names:
                continue
            
            if has_header:
                df = pd.read_excel(xlsx, sheet_name=sheet_name)
            else:
                df = pd.read_excel(xlsx, sheet_name=sheet_name, header=None)
                df.columns = [f"Column {i+1}" for i in range(len(df.columns))]
            
            sheet_selected_cols = selected_columns_map.get(str(sheet_name), [])
            if sheet_selected_cols:
                existing_cols = [c for c in sheet_selected_cols if c in df.columns]
                if existing_cols:
                    df = df[existing_cols]
            
            if max_rows and max_rows > 0:
                df = df.head(max_rows)
            
            df = df.where(pd.notnull(df), None)
            tables.append(self._df_to_table(df, id_=str(sheet_name), name=str(sheet_name)))
        
        return tables
    
    # ─── Smart mode extraction ───────────────────────────────────────
    
    def _extract_regions(
        self,
        xlsx,
        regions: list[dict[str, Any]],
        has_header: bool,
        max_rows: int | None,
    ) -> list[TableData]:
        """Extract specific rectangular regions detected by ExcelTableDetector.
        
        Каждый region содержит:
          sheet_name, start_row, start_col, end_row, end_col,
          header_row (1-based | None), table_name, selected_columns
        """
        import openpyxl
        import pandas as pd
        
        tables: list[TableData] = []
        
        # Cache loaded sheets
        wb = xlsx.book if hasattr(xlsx, 'book') else openpyxl.load_workbook(xlsx.io, data_only=True)
        
        for region in regions:
            sheet_name = region["sheet_name"]
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            sr, sc = region["start_row"], region["start_col"]
            er, ec = region["end_row"], region["end_col"]
            header_row = region.get("header_row")
            table_name = region.get("table_name", f"{sheet_name}_{sr}_{sc}")
            selected_cols = region.get("selected_columns", [])
            
            # Build column names from header row or generate
            col_names: list[str] = []
            data_start_row = sr
            
            if header_row and has_header:
                for c in range(sc, ec + 1):
                    val = ws.cell(row=header_row, column=c).value
                    col_names.append(str(val) if val is not None else f"Column {c - sc + 1}")
                data_start_row = header_row + 1
            else:
                for c in range(sc, ec + 1):
                    col_names.append(f"Column {c - sc + 1}")
            
            # Read data rows
            rows_data: list[dict[str, Any]] = []
            row_limit = max_rows if (max_rows and max_rows > 0) else (er - data_start_row + 1)
            actual_end = min(er, data_start_row + row_limit - 1)
            
            for r in range(data_start_row, actual_end + 1):
                row_dict: dict[str, Any] = {}
                for ci, c in enumerate(range(sc, ec + 1)):
                    val = ws.cell(row=r, column=c).value
                    row_dict[col_names[ci]] = val
                rows_data.append(row_dict)
            
            # Convert to DataFrame for type inference + filtering
            df = pd.DataFrame(rows_data, columns=col_names)
            
            # Filter columns
            if selected_cols:
                existing = [c for c in selected_cols if c in df.columns]
                if existing:
                    df = df[existing]
            
            df = df.where(pd.notnull(df), None)
            
            table_id = f"{sheet_name}:{sr},{sc}-{er},{ec}"
            display_name = table_name or table_id
            tables.append(self._df_to_table(df, id_=table_id, name=display_name))
        
        return tables
    
    # ─── Helpers ─────────────────────────────────────────────────────
    
    def _df_to_table(self, df, id_: str, name: str) -> TableData:
        """Convert a pandas DataFrame to TableData."""
        import pandas as pd
        
        columns = [
            {"name": str(col), "type": self._infer_pandas_type(df[col])}
            for col in df.columns
        ]
        
        rows = []
        for _, row in df.iterrows():
            row_dict = {}
            for col in df.columns:
                val = row[col]
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    row_dict[str(col)] = None
                else:
                    row_dict[str(col)] = val
            rows.append(row_dict)
        
        return TableData(id=id_, name=name, columns=columns, rows=rows)
    
    def _infer_pandas_type(self, series) -> str:
        """Infer column type from pandas series."""
        import pandas as pd
        
        if pd.api.types.is_numeric_dtype(series):
            return "number"
        elif pd.api.types.is_datetime64_any_dtype(series):
            return "date"
        return "text"
    
    def get_dialog_schema(self) -> dict[str, Any]:
        """Get JSON Schema for Excel dialog."""
        return {
            "type": "object",
            "properties": {
                "file": {
                    "type": "file",
                    "accept": ".xlsx,.xls",
                    "title": "Excel файл",
                },
                "extraction_code": {
                    "type": "string",
                    "format": "python",
                    "title": "Python код извлечения",
                },
            },
        }
