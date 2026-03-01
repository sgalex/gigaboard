"""Excel Table Detector - эвристический + AI поиск таблиц в произвольных Excel файлах.

Проблема: Данные в Excel могут быть размещены произвольно — таблица не обязательно
начинается с A1, на одном листе может быть несколько таблиц, между ними — текст,
пустые строки, заголовки отчётов, merged cells.

Решение: Гибридный подход:
1. Эвристический pass — openpyxl raw cells → нахождение прямоугольных регионов
2. AI pass (GigaChat/StructurizerAgent) — уточнение границ при неоднозначности

См. docs/SOURCE_NODE_CONCEPT_V2.md — Excel Dialog.
"""
import logging
import io
from dataclasses import dataclass, field
from typing import Any

logger = logging.getLogger(__name__)


@dataclass
class DetectedRegion:
    """Обнаруженный прямоугольный регион с данными на листе."""
    sheet_name: str
    start_row: int  # 1-based
    start_col: int  # 1-based
    end_row: int    # 1-based, inclusive
    end_col: int    # 1-based, inclusive
    header_row: int | None = None  # 1-based row that contains headers (None = no header)
    confidence: float = 0.0  # 0.0..1.0 — уверенность в правильности региона
    table_name: str = ""  # Имя таблицы (определённое по заголовку или позиции)
    columns: list[dict[str, str]] = field(default_factory=list)  # [{name, type}]
    preview_rows: list[dict[str, Any]] = field(default_factory=list)  # first N rows as dicts
    row_count: int = 0  # количество строк данных (без заголовка)
    detection_method: str = "heuristic"  # heuristic | ai | hybrid

    @property
    def col_letter_start(self) -> str:
        """Excel column letter for start_col (1-based)."""
        return _col_to_letter(self.start_col)

    @property
    def col_letter_end(self) -> str:
        """Excel column letter for end_col (1-based)."""
        return _col_to_letter(self.end_col)

    @property
    def range_str(self) -> str:
        """Excel-style range string, e.g. 'A1:D15'."""
        return f"{self.col_letter_start}{self.start_row}:{self.col_letter_end}{self.end_row}"

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "start_row": self.start_row,
            "start_col": self.start_col,
            "end_row": self.end_row,
            "end_col": self.end_col,
            "header_row": self.header_row,
            "confidence": self.confidence,
            "table_name": self.table_name,
            "columns": self.columns,
            "preview_rows": self.preview_rows,
            "row_count": self.row_count,
            "range_str": self.range_str,
            "detection_method": self.detection_method,
        }


@dataclass
class SheetDetectionResult:
    """Результат детекции для одного листа."""
    sheet_name: str
    total_rows: int
    total_cols: int
    regions: list[DetectedRegion] = field(default_factory=list)
    # Grid snapshot для визуализации на фронтенде
    # Матрица типов ячеек: 'e'=empty, 'd'=data, 'h'=header-candidate, 'm'=merged
    grid_map: list[list[str]] = field(default_factory=list)
    grid_rows: int = 0
    grid_cols: int = 0

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "total_rows": self.total_rows,
            "total_cols": self.total_cols,
            "regions": [r.to_dict() for r in self.regions],
            "grid_map": self.grid_map,
            "grid_rows": self.grid_rows,
            "grid_cols": self.grid_cols,
        }


@dataclass
class ExcelDetectionResult:
    """Полный результат детекции для всего файла."""
    sheet_names: list[str]
    sheets: list[SheetDetectionResult] = field(default_factory=list)
    total_tables_found: int = 0
    detection_method: str = "heuristic"  # heuristic | ai | hybrid

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_names": self.sheet_names,
            "sheets": [s.to_dict() for s in self.sheets],
            "total_tables_found": self.total_tables_found,
            "detection_method": self.detection_method,
        }


# ─── Основной детектор ──────────────────────────────────────────────

class ExcelTableDetector:
    """Гибридный детектор таблиц в Excel файлах.
    
    Алгоритм:
    1. Через openpyxl читаем raw-ячейки (без pandas — сохраняем позиции)
    2. Строим матрицу заполненности (grid_map)
    3. Находим connected components — прямоугольные регионы с данными
    4. Для каждого региона определяем header row и типы столбцов
    5. Если confidence < threshold — запускаем AI уточнение (GigaChat)
    """

    # Минимальное количество заполненных ячеек в строке для «якорной строки»
    MIN_CELLS_IN_ROW = 2
    # Минимальное количество строк данных для таблицы
    MIN_DATA_ROWS = 1
    # Максимальное количество пустых строк внутри таблицы перед разрывом
    MAX_EMPTY_GAP = 2
    # Максимальный размер grid snapshot для визуализации (строк × столбцов)
    MAX_GRID_ROWS = 60
    MAX_GRID_COLS = 30
    # Количество preview rows
    PREVIEW_ROWS = 10
    # Порог confidence для AI уточнения
    AI_CONFIDENCE_THRESHOLD = 0.5

    def __init__(self, gigachat_service=None):
        """
        Args:
            gigachat_service: Опциональный GigaChatService для AI уточнения.
                             Если None — только эвристика.
        """
        self.gigachat = gigachat_service

    async def detect(self, file_content: bytes) -> ExcelDetectionResult:
        """Основная точка входа: определяет таблицы во всём Excel файле."""
        import openpyxl

        excel_file = io.BytesIO(file_content)
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

        sheet_names = wb.sheetnames
        sheet_results: list[SheetDetectionResult] = []
        total_tables = 0

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            result = await self._detect_sheet(ws, sheet_name)
            sheet_results.append(result)
            total_tables += len(result.regions)

        wb.close()

        # Определяем итоговый метод (hybrid если хотя бы один регион получил AI)
        methods = set()
        for sr in sheet_results:
            for r in sr.regions:
                methods.add(r.detection_method)
        detection_method = "hybrid" if len(methods) > 1 else (methods.pop() if methods else "heuristic")

        return ExcelDetectionResult(
            sheet_names=sheet_names,
            sheets=sheet_results,
            total_tables_found=total_tables,
            detection_method=detection_method,
        )

    async def _detect_sheet(self, ws, sheet_name: str) -> SheetDetectionResult:
        """Обнаруживает таблицы на одном листе."""
        # 1. Читаем все ячейки в матрицу
        cells, max_row, max_col = self._read_cells(ws)

        if max_row == 0 or max_col == 0:
            return SheetDetectionResult(
                sheet_name=sheet_name,
                total_rows=0,
                total_cols=0,
            )

        # 2. Строим grid_map для визуализации
        grid_map = self._build_grid_map(cells, max_row, max_col)

        # 3. Находим прямоугольные регионы
        regions = self._find_regions(cells, max_row, max_col, sheet_name)

        # 4. Для каждого региона определяем header и типы
        for region in regions:
            self._analyze_region(region, cells)

        # 5. AI уточнение для низко-confidence регионов
        low_confidence = [r for r in regions if r.confidence < self.AI_CONFIDENCE_THRESHOLD]
        if low_confidence and self.gigachat:
            await self._ai_refine(cells, max_row, max_col, sheet_name, regions)

        # Ограничиваем grid_map для фронтенда
        visible_rows = min(max_row, self.MAX_GRID_ROWS)
        visible_cols = min(max_col, self.MAX_GRID_COLS)
        trimmed_grid = [row[:visible_cols] for row in grid_map[:visible_rows]]

        return SheetDetectionResult(
            sheet_name=sheet_name,
            total_rows=max_row,
            total_cols=max_col,
            regions=regions,
            grid_map=trimmed_grid,
            grid_rows=visible_rows,
            grid_cols=visible_cols,
        )

    # ─── Шаг 1: Чтение raw-ячеек ──────────────────────────────────────

    def _read_cells(self, ws) -> tuple[dict[tuple[int, int], Any], int, int]:
        """Читает все ячейки листа в dict {(row, col): value}.
        
        Returns:
            (cells_dict, max_row, max_col) — 1-based координаты.
        """
        cells: dict[tuple[int, int], Any] = {}
        max_row = 0
        max_col = 0

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    r, c = cell.row, cell.column
                    cells[(r, c)] = cell.value
                    max_row = max(max_row, r)
                    max_col = max(max_col, c)

        return cells, max_row, max_col

    # ─── Шаг 2: Grid map для визуализации ──────────────────────────────

    def _build_grid_map(
        self,
        cells: dict[tuple[int, int], Any],
        max_row: int,
        max_col: int,
    ) -> list[list[str]]:
        """Строит матрицу типов ячеек для мини-карты фронтенда.
        
        Типы: 'e' = empty, 'd' = data (number/date), 't' = text, 'h' = header-candidate
        """
        rows = min(max_row, self.MAX_GRID_ROWS)
        cols = min(max_col, self.MAX_GRID_COLS)

        grid: list[list[str]] = []
        for r in range(1, rows + 1):
            row_types: list[str] = []
            for c in range(1, cols + 1):
                val = cells.get((r, c))
                if val is None:
                    row_types.append("e")
                elif isinstance(val, (int, float)):
                    row_types.append("d")
                elif isinstance(val, str):
                    # Heuristic: if a text cell has neighbors below with numbers → likely header
                    below_val = cells.get((r + 1, c))
                    if below_val is not None and isinstance(below_val, (int, float)):
                        row_types.append("h")
                    else:
                        row_types.append("t")
                else:
                    row_types.append("d")  # datetime и др.
            grid.append(row_types)

        return grid

    # ─── Шаг 3: Поиск прямоугольных регионов ──────────────────────────

    def _find_regions(
        self,
        cells: dict[tuple[int, int], Any],
        max_row: int,
        max_col: int,
        sheet_name: str,
    ) -> list[DetectedRegion]:
        """Находит связные прямоугольные регионы данных.
        
        Алгоритм:
        1. Построить карту занятости строк (row → set of filled columns)
        2. Найти «якорные строки» с ≥ MIN_CELLS_IN_ROW заполненных ячеек
        3. Группировать смежные якорные строки с одинаковым column span
        4. Каждая группа = кандидат на таблицу
        """
        # row_profile: для каждой строки — множество заполненных столбцов
        row_profile: dict[int, set[int]] = {}
        for (r, c), val in cells.items():
            if val is not None:
                if r not in row_profile:
                    row_profile[r] = set()
                row_profile[r].add(c)

        if not row_profile:
            return []

        # Ищем якорные строки
        anchor_rows = []
        for r in range(1, max_row + 1):
            cols_in_row = row_profile.get(r, set())
            if len(cols_in_row) >= self.MIN_CELLS_IN_ROW:
                anchor_rows.append(r)

        if not anchor_rows:
            return []

        # Группируем смежные якорные строки
        # Допускаем gap до MAX_EMPTY_GAP пустых строк внутри таблицы
        groups: list[list[int]] = []
        current_group: list[int] = [anchor_rows[0]]

        for i in range(1, len(anchor_rows)):
            gap = anchor_rows[i] - anchor_rows[i - 1] - 1
            # Считаем перекрытие столбцов между группами
            prev_cols = row_profile.get(anchor_rows[i - 1], set())
            curr_cols = row_profile.get(anchor_rows[i], set())
            overlap = len(prev_cols & curr_cols)
            total = max(len(prev_cols | curr_cols), 1)
            overlap_ratio = overlap / total

            if gap <= self.MAX_EMPTY_GAP and overlap_ratio >= 0.3:
                current_group.append(anchor_rows[i])
            else:
                groups.append(current_group)
                current_group = [anchor_rows[i]]

        groups.append(current_group)

        # Для каждой группы определяем bounding box
        regions: list[DetectedRegion] = []
        table_idx = 0

        for group in groups:
            if len(group) < self.MIN_DATA_ROWS:
                continue

            # Объединяем все столбцы в группе
            all_cols: set[int] = set()
            for r in group:
                all_cols.update(row_profile.get(r, set()))

            if len(all_cols) < self.MIN_CELLS_IN_ROW:
                continue

            start_row = group[0]
            end_row = group[-1]
            start_col = min(all_cols)
            end_col = max(all_cols)

            # Расширяем вниз: ищем строки ниже end_row, которые ещё содержат данные
            # в том же диапазоне столбцов
            for r in range(end_row + 1, min(end_row + 100, max_row + 1)):
                cols_in_row = row_profile.get(r, set())
                in_range = cols_in_row & set(range(start_col, end_col + 1))
                if len(in_range) >= max(1, len(all_cols) * 0.3):
                    end_row = r
                else:
                    # Допускаем gap
                    gap_ahead = 0
                    found_more = False
                    for r2 in range(r + 1, min(r + self.MAX_EMPTY_GAP + 1, max_row + 1)):
                        c2 = row_profile.get(r2, set()) & set(range(start_col, end_col + 1))
                        if len(c2) >= max(1, len(all_cols) * 0.3):
                            found_more = True
                            end_row = r2
                            break
                    if not found_more:
                        break

            table_idx += 1
            regions.append(DetectedRegion(
                sheet_name=sheet_name,
                start_row=start_row,
                start_col=start_col,
                end_row=end_row,
                end_col=end_col,
                table_name=f"{sheet_name}_table_{table_idx}",
            ))

        return regions

    # ─── Шаг 4: Анализ каждого региона ─────────────────────────────────

    def _analyze_region(self, region: DetectedRegion, cells: dict[tuple[int, int], Any]):
        """Определяет header row, типы столбцов и строит preview."""
        sr, sc, er, ec = region.start_row, region.start_col, region.end_row, region.end_col

        # Определяем header: первая строка, где большинство ячеек — текст
        first_row_vals = [cells.get((sr, c)) for c in range(sc, ec + 1)]
        second_row_vals = [cells.get((sr + 1, c)) for c in range(sc, ec + 1)] if sr < er else []

        first_row_text_ratio = _text_ratio(first_row_vals)
        second_row_text_ratio = _text_ratio(second_row_vals) if second_row_vals else 1.0

        # Если первая строка преимущественно текстовая, а вторая — нет → header
        has_header = (
            first_row_text_ratio >= 0.6
            and (second_row_text_ratio < first_row_text_ratio or not second_row_vals)
        )

        if has_header:
            region.header_row = sr
            data_start = sr + 1
            # Имена столбцов из header
            col_names = []
            for c in range(sc, ec + 1):
                val = cells.get((sr, c))
                col_names.append(str(val).strip() if val is not None else f"Column_{c}")
        else:
            region.header_row = None
            data_start = sr
            col_names = [f"Column_{c}" for c in range(sc, ec + 1)]

        # Определяем типы столбцов по данным
        columns: list[dict[str, str]] = []
        for idx, c in enumerate(range(sc, ec + 1)):
            col_vals = [cells.get((r, c)) for r in range(data_start, min(data_start + 20, er + 1))]
            col_vals = [v for v in col_vals if v is not None]
            col_type = _infer_type(col_vals)
            columns.append({"name": col_names[idx], "type": col_type})

        region.columns = columns
        region.row_count = er - data_start + 1

        # Confidence scoring
        confidence = 0.0
        # Больше строк → больше уверенности
        if region.row_count >= 5:
            confidence += 0.3
        elif region.row_count >= 2:
            confidence += 0.15
        # Есть header → +0.3
        if has_header:
            confidence += 0.3
        # Консистентные типы столбцов → +0.2
        consistent_cols = sum(1 for col in columns if col["type"] != "mixed")
        if columns:
            confidence += 0.2 * (consistent_cols / len(columns))
        # Более 2 столбцов → +0.2
        if len(columns) >= 3:
            confidence += 0.2

        region.confidence = min(confidence, 1.0)

        # Preview rows
        preview_rows: list[dict[str, Any]] = []
        for r in range(data_start, min(data_start + self.PREVIEW_ROWS, er + 1)):
            row_dict: dict[str, Any] = {}
            for idx, c in enumerate(range(sc, ec + 1)):
                val = cells.get((r, c))
                name = columns[idx]["name"]
                row_dict[name] = _safe_json_value(val)
            preview_rows.append(row_dict)

        region.preview_rows = preview_rows

        # Попытка дать осмысленное имя
        if has_header and col_names:
            # Используем первые 2-3 столбца как подсказку
            name_hint = ", ".join(col_names[:3])
            region.table_name = f"{region.sheet_name} ({name_hint}...)"

    # ─── Шаг 5: AI уточнение через GigaChat ──────────────────────────

    async def _ai_refine(
        self,
        cells: dict[tuple[int, int], Any],
        max_row: int,
        max_col: int,
        sheet_name: str,
        regions: list[DetectedRegion],
    ):
        """Используем GigaChat для уточнения найденных регионов.
        
        Отправляем текстовое представление листа + текущие найденные регионы
        и просим AI скорректировать или подтвердить.
        """
        if not self.gigachat:
            return

        try:
            # Формируем grid snapshot (CSV-like)
            snapshot = self._build_text_snapshot(cells, max_row, max_col)
            
            # Описание найденных регионов
            regions_desc = "\n".join([
                f"  - Регион {i+1}: {r.range_str} (confidence={r.confidence:.2f}, "
                f"header_row={r.header_row}, cols={len(r.columns)}, rows={r.row_count})"
                for i, r in enumerate(regions)
            ])

            # System prompt для задачи
            system_prompt = """Ты — эксперт по анализу структуры Excel файлов.
Тебе даётся текстовое представление содержимого листа Excel и список предварительно
найденных таблиц (регионов данных). Твоя задача — уточнить или исправить их.

Верни ТОЛЬКО чистый JSON (без markdown-обёрток):
{
  "regions": [
    {
      "start_row": 1,
      "start_col": 1,
      "end_row": 10,
      "end_col": 5,
      "header_row": 1,
      "table_name": "описательное имя таблицы",
      "confidence": 0.95,
      "notes": "почему этот регион — таблица"
    }
  ],
  "notes": "общий комментарий"
}

Правила:
- Координаты 1-based (строки и столбцы начинаются с 1)
- header_row = null если нет заголовка
- Не включай декоративные элементы (заголовки отчётов, итоги) в данные таблицы
- Если таблица начинается не с A1, точно укажи координаты
- Объединяй регионы если они — одна таблица с пустой строкой
- Разделяй регионы если в одном найдено несколько таблиц"""

            user_prompt = f"""Лист Excel: "{sheet_name}"

=== СОДЕРЖИМОЕ ЛИСТА (grid snapshot) ===
{snapshot}

=== ПРЕДВАРИТЕЛЬНО НАЙДЕННЫЕ РЕГИОНЫ ===
{regions_desc if regions else "Регионы не найдены эвристикой."}

Проанализируй содержимое листа и уточни найденные регионы (или найди новые)."""

            messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ]

            response = await self.gigachat.chat_completion(
                messages=messages,
                temperature=0.2,
                max_tokens=3000,
            )

            # Парсим ответ
            import json
            import re as re_mod
            
            response = response.strip()
            # Ищем JSON в markdown блоке
            json_match = re_mod.search(r'```(?:json)?\s*(\{.*?\})\s*```', response, re_mod.DOTALL)
            if json_match:
                response = json_match.group(1)
            # Ищем JSON объект
            json_match = re_mod.search(r'\{.*\}', response, re_mod.DOTALL)
            if json_match:
                ai_result = json.loads(json_match.group())
            else:
                logger.warning("AI refine: could not parse JSON response")
                return

            # Обновляем регионы на основе AI ответа
            ai_regions = ai_result.get("regions", [])
            if not ai_regions:
                return

            # Заменяем регионы AI-версиями
            new_regions: list[DetectedRegion] = []
            for ai_reg in ai_regions:
                sr = ai_reg.get("start_row", 1)
                sc = ai_reg.get("start_col", 1)
                er = ai_reg.get("end_row", max_row)
                ec = ai_reg.get("end_col", max_col)
                hr = ai_reg.get("header_row")
                conf = ai_reg.get("confidence", 0.8)
                name = ai_reg.get("table_name", f"{sheet_name}_table")

                region = DetectedRegion(
                    sheet_name=sheet_name,
                    start_row=sr,
                    start_col=sc,
                    end_row=er,
                    end_col=ec,
                    header_row=hr,
                    confidence=conf,
                    table_name=name,
                    detection_method="ai",
                )
                # Заполняем columns и preview_rows
                self._analyze_region(region, cells)
                # Берём confidence от AI если выше
                region.confidence = max(region.confidence, conf)
                region.detection_method = "ai"
                new_regions.append(region)

            # Заменяем исходные регионы
            regions.clear()
            regions.extend(new_regions)

        except Exception as e:
            logger.error(f"AI refine failed: {e}", exc_info=True)
            # Оставляем эвристические результаты как есть

    def _build_text_snapshot(
        self,
        cells: dict[tuple[int, int], Any],
        max_row: int,
        max_col: int,
    ) -> str:
        """Строит текстовое CSV-подобное представление листа для AI."""
        snapshot_rows = min(max_row, 50)  # Ограничиваем для контекста LLM
        snapshot_cols = min(max_col, 20)
        
        lines = []
        # Заголовок с номерами столбцов
        header = "     | " + " | ".join(
            f"{_col_to_letter(c):>8}" for c in range(1, snapshot_cols + 1)
        )
        lines.append(header)
        lines.append("─" * len(header))

        for r in range(1, snapshot_rows + 1):
            vals = []
            for c in range(1, snapshot_cols + 1):
                val = cells.get((r, c))
                if val is None:
                    vals.append("        ")  # 8 chars empty
                else:
                    s = str(val)[:8].ljust(8)
                    vals.append(s)
            line = f"{r:4d} | " + " | ".join(vals)
            lines.append(line)

        if max_row > snapshot_rows:
            lines.append(f"... ещё {max_row - snapshot_rows} строк")
        if max_col > snapshot_cols:
            lines.append(f"... ещё {max_col - snapshot_cols} столбцов")

        return "\n".join(lines)


# ─── Utility functions ──────────────────────────────────────────────

def _col_to_letter(col: int) -> str:
    """Конвертирует номер столбца (1-based) в Excel-букву: 1→A, 26→Z, 27→AA."""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _text_ratio(values: list[Any]) -> float:
    """Доля текстовых значений в списке (ignoring None)."""
    non_none = [v for v in values if v is not None]
    if not non_none:
        return 0.0
    text_count = sum(1 for v in non_none if isinstance(v, str))
    return text_count / len(non_none)


def _infer_type(values: list[Any]) -> str:
    """Определяет тип столбца по значениям."""
    if not values:
        return "text"

    import datetime

    type_counts = {"number": 0, "text": 0, "date": 0, "bool": 0}
    for v in values:
        if isinstance(v, bool):
            type_counts["bool"] += 1
        elif isinstance(v, (int, float)):
            type_counts["number"] += 1
        elif isinstance(v, datetime.datetime):
            type_counts["date"] += 1
        elif isinstance(v, str):
            # Попытка распознать число в строке
            stripped = v.strip().replace(",", ".").replace(" ", "")
            try:
                float(stripped)
                type_counts["number"] += 1
                continue
            except ValueError:
                pass
            type_counts["text"] += 1
        else:
            type_counts["text"] += 1

    # Находим доминирующий тип (>50%)
    total = sum(type_counts.values())
    if total == 0:
        return "text"

    dominant = max(type_counts, key=type_counts.get)  # type: ignore
    if type_counts[dominant] / total >= 0.5:
        return dominant

    return "mixed"


def _safe_json_value(val: Any) -> Any:
    """Конвертирует значение в JSON-safe формат."""
    if val is None:
        return None
    import datetime
    if isinstance(val, datetime.datetime):
        return val.isoformat()
    if isinstance(val, (int, float, bool, str)):
        return val
    return str(val)
