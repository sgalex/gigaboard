"""File upload routes."""
import logging
import csv
import io
from typing import Annotated
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel

from app.core import get_db
from app.core.config import settings
from app.models import User
from app.routes.auth import get_current_user
from app.services.file_storage import get_storage

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/files", tags=["files"])

# Max file size from settings
MAX_FILE_SIZE = settings.STORAGE_MAX_FILE_SIZE_MB * 1024 * 1024


@router.post("/upload")
async def upload_file(
    file: Annotated[UploadFile, File()],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Upload a file and return file metadata.
    
    Supports local filesystem or S3-compatible storage (MinIO, AWS S3, Yandex Object Storage).
    Storage backend configured via STORAGE_BACKEND env variable.
    
    Returns:
        {
            "file_id": "uuid",
            "filename": "original_name.pdf",
            "mime_type": "application/pdf",
            "size_bytes": 12345,
            "storage_path": "user_id/2026/01/uuid.pdf"
        }
    """
    try:
        # Validate file size by reading chunks
        file_size = 0
        chunks = []
        
        while chunk := await file.read(8192):
            file_size += len(chunk)
            if file_size > MAX_FILE_SIZE:
                raise HTTPException(
                    status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                    detail=f"File too large. Max size: {settings.STORAGE_MAX_FILE_SIZE_MB}MB"
                )
            chunks.append(chunk)
        
        # Generate unique file ID
        file_id = str(uuid4())
        
        # Save file using storage service
        storage = get_storage()
        
        # Create a file-like object from chunks
        import io
        file_data = io.BytesIO(b"".join(chunks))
        
        storage_path = await storage.save(
            file_id=file_id,
            file_data=file_data,
            user_id=current_user.id,  # type: ignore
            filename=file.filename or "unknown",
            mime_type=file.content_type or "application/octet-stream",
            db=db
        )
        
        # Commit transaction to persist file in database
        await db.commit()
        
        logger.info(f"File uploaded: {file.filename} ({file_size} bytes) by user {current_user.id}")
        
        return {
            "file_id": file_id,
            "filename": file.filename or "unknown",
            "mime_type": file.content_type or "application/octet-stream",
            "size_bytes": file_size,
            "storage_path": storage_path
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"File upload failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"File upload failed: {str(e)}"
        )


@router.get("/download/{file_id}")
async def download_file(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Download a file by ID.
    
    For database storage: returns file directly via Response with bytes
    For local storage: returns file directly via FileResponse
    For S3 storage: returns presigned URL (redirect)
    """
    try:
        storage = get_storage()
        
        # Check if file exists (pass db for DatabaseFileStorage)
        if not await storage.exists(file_id, db=db):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found"
            )
        
        # Get file content/path/URL (pass db for DatabaseFileStorage)
        file_location = await storage.get(file_id, db=db)
        
        # For database storage, file_location is bytes
        if isinstance(file_location, bytes):
            from fastapi.responses import Response
            from app.models import UploadedFile
            from sqlalchemy import select
            from uuid import UUID
            
            # Get file metadata for proper response
            result = await db.execute(
                select(UploadedFile).where(UploadedFile.id == UUID(file_id))
            )
            uploaded_file = result.scalar_one_or_none()
            
            return Response(
                content=file_location,
                media_type=uploaded_file.mime_type if uploaded_file else "application/octet-stream",
                headers={
                    "Content-Disposition": f'attachment; filename="{uploaded_file.filename if uploaded_file else file_id}"'
                }
            )
        
        # For S3 storage, file_location is a presigned URL
        if settings.STORAGE_BACKEND == "s3":
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url=str(file_location))
        
        # For local storage, file_location is a Path object
        from pathlib import Path
        if isinstance(file_location, (str, Path)):
            return FileResponse(
                path=str(file_location),
                filename=Path(file_location).name,
                media_type="application/octet-stream"
            )
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Invalid storage response"
        )
        
    except HTTPException:
        raise
    except FileNotFoundError:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="File not found"
        )
    except Exception as e:
        logger.error(f"File download failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"File download failed: {str(e)}"
        )


# ---------- Public image preview (no auth) ----------
ALLOWED_IMAGE_MIMES = {"image/png", "image/jpeg", "image/gif", "image/webp", "image/svg+xml", "image/avif"}


@router.get("/image/{file_id}")
async def get_image(
    file_id: str,
    db: AsyncSession = Depends(get_db),
):
    """Serve an uploaded image publicly (no auth required).

    Only files with image/* MIME type are served.
    Works with all storage backends (database, local, s3).
    Returns inline with Cache-Control headers.
    """
    from fastapi.responses import Response
    from app.models import UploadedFile as UF
    from sqlalchemy import select
    from uuid import UUID

    try:
        # Get file metadata
        result = await db.execute(
            select(UF).where(UF.id == UUID(file_id))
        )
        uploaded_file = result.scalar_one_or_none()

        if not uploaded_file:
            raise HTTPException(status_code=404, detail="File not found")

        # If mime_type is generic, try to guess from filename
        mime_type = uploaded_file.mime_type
        if mime_type == "application/octet-stream" and uploaded_file.filename:
            import mimetypes
            guessed, _ = mimetypes.guess_type(uploaded_file.filename)
            if guessed:
                mime_type = guessed

        if mime_type not in ALLOWED_IMAGE_MIMES:
            raise HTTPException(status_code=403, detail="Not an image file")

        # Get file content via storage backend
        storage = get_storage()
        file_content = await storage.get(file_id, db=db)

        # For database storage, file_content is bytes
        if isinstance(file_content, bytes):
            return Response(
                content=file_content,
                media_type=mime_type,
                headers={
                    "Content-Disposition": f'inline; filename="{uploaded_file.filename}"',
                    "Cache-Control": "public, max-age=86400",
                },
            )

        # For local storage, file_content is a Path
        from pathlib import Path
        if isinstance(file_content, (str, Path)):
            return FileResponse(
                path=str(file_content),
                media_type=mime_type,
                filename=uploaded_file.filename,
                headers={"Cache-Control": "public, max-age=86400"},
            )

        # For S3 storage, file_content is a presigned URL
        if settings.STORAGE_BACKEND == "s3":
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url=str(file_content))

        raise HTTPException(status_code=500, detail="Invalid storage response")
    except HTTPException:
        raise
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found")
    except Exception as e:
        logger.error(f"Image serve failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Image serve failed")


# CSV Analysis Response Schema
class CSVAnalysisResult(BaseModel):
    """CSV analysis result."""
    delimiter: str
    encoding: str
    has_header: bool
    rows_count: int
    columns: list[dict]
    preview_rows: list[dict]


@router.post("/{file_id}/analyze-csv", response_model=CSVAnalysisResult)
async def analyze_csv_file(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Analyze CSV file and detect encoding, delimiter, columns.
    
    Uses Python csv.Sniffer for delimiter detection and chardet for encoding.
    Returns structure similar to frontend analysis but more accurate.
    """
    try:
        storage = get_storage()
        
        # Download file content
        file_content = await storage.get(
            file_id=file_id,
            db=db
        )
        
        # Detect encoding using chardet
        try:
            import chardet
            detected = chardet.detect(file_content[:100000])
            encoding = detected['encoding'] or 'utf-8'
            confidence = detected['confidence']
            
            # If confidence is low, try common encodings
            if confidence < 0.7:
                for fallback_enc in ['utf-8', 'windows-1251', 'cp1251', 'latin1']:
                    try:
                        file_content.decode(fallback_enc)
                        encoding = fallback_enc
                        break
                    except UnicodeDecodeError:
                        continue
        except ImportError:
            # chardet not installed, fallback logic
            encoding = 'utf-8'
            try:
                file_content.decode('utf-8')
            except UnicodeDecodeError:
                encoding = 'windows-1251'
        
        # Decode content
        try:
            content_str = file_content.decode(encoding)
        except UnicodeDecodeError:
            # Last resort
            content_str = file_content.decode('utf-8', errors='replace')
            encoding = 'utf-8'
        
        # Detect delimiter using csv.Sniffer
        sniffer = csv.Sniffer()
        sample = content_str[:4096]
        
        try:
            dialect = sniffer.sniff(sample, delimiters=',;\t|')
            delimiter = dialect.delimiter
        except csv.Error:
            # Fallback: count occurrences
            delimiters = {',': 0, ';': 0, '\t': 0, '|': 0}
            first_line = content_str.split('\n')[0] if content_str else ''
            for delim in delimiters:
                delimiters[delim] = first_line.count(delim)
            delimiter = max(delimiters, key=delimiters.get)
        
        # Detect if first row is header
        lines = content_str.split('\n')
        lines = [l for l in lines if l.strip()]
        
        if len(lines) < 2:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV file is too small to analyze"
            )
        
        # Try to detect header using Sniffer
        try:
            has_header = sniffer.has_header(sample)
        except csv.Error:
            has_header = True  # Assume header by default
        
        # Parse CSV
        reader = csv.reader(io.StringIO(content_str), delimiter=delimiter)
        rows_list = list(reader)
        
        if not rows_list:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV file is empty"
            )
        
        # Extract headers and data
        if has_header:
            headers = rows_list[0]
            data_rows = rows_list[1:min(11, len(rows_list))]  # First 10 data rows
        else:
            headers = [f"Column {i+1}" for i in range(len(rows_list[0]))]
            data_rows = rows_list[:10]
        
        # Clean headers
        headers = [h.strip().replace('"', '').replace("'", '') for h in headers]
        
        # Build preview rows
        preview_rows = []
        for row in data_rows:
            row_dict = {}
            for i, header in enumerate(headers):
                row_dict[header] = row[i] if i < len(row) else ''
            preview_rows.append(row_dict)
        
        # Detect column types
        columns = []
        for i, header in enumerate(headers):
            samples = [row[i] for row in data_rows if i < len(row) and row[i].strip()]
            
            # Type detection
            is_numeric = all(val.replace('.', '', 1).replace('-', '', 1).isdigit() for val in samples if val)
            is_date = any(
                val.count('-') == 2 and len(val.split('-')[0]) == 4 
                for val in samples if val
            )
            
            col_type = 'дата' if is_date else ('число' if is_numeric else 'текст')
            
            columns.append({
                'name': header,
                'type': col_type,
                'sample_values': samples[:3]
            })
        
        # Convert delimiter back to display format
        delimiter_display = 'tab' if delimiter == '\t' else delimiter
        
        return CSVAnalysisResult(
            delimiter=delimiter_display,
            encoding=encoding,
            has_header=has_header,
            rows_count=len(rows_list) - (1 if has_header else 0),
            columns=columns,
            preview_rows=preview_rows
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"CSV analysis failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"CSV analysis failed: {str(e)}"
        )


# ============================================
# Excel Analysis
# ============================================

class ExcelSheetInfo(BaseModel):
    """Information about a single Excel sheet."""
    name: str
    rows_count: int
    columns: list[dict]  # [{name, type, sample_values}]
    preview_rows: list[dict]


class ExcelAnalysisResult(BaseModel):
    """Excel analysis result."""
    sheet_names: list[str]
    sheets: list[ExcelSheetInfo]
    total_rows: int


@router.post("/{file_id}/analyze-excel", response_model=ExcelAnalysisResult)
async def analyze_excel_file(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Analyze Excel file: detect sheets, columns, types, preview data.
    
    Returns sheet-by-sheet analysis with column types and preview rows.
    See docs/SOURCE_NODE_CONCEPT_V2.md - section "Excel Dialog".
    """
    try:
        import pandas as pd

        storage = get_storage()
        file_content = await storage.get(file_id=file_id, db=db)

        if file_content is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found"
            )

        # Read Excel file
        excel_file = io.BytesIO(file_content)
        xlsx = pd.ExcelFile(excel_file)

        sheet_names = xlsx.sheet_names
        sheets: list[ExcelSheetInfo] = []
        total_rows = 0

        for sheet_name in sheet_names:
            df = pd.read_excel(xlsx, sheet_name=sheet_name)
            total_rows += len(df)

            # Detect column types
            columns = []
            for col_name in df.columns:
                col = df[col_name]
                samples = col.dropna().head(5).astype(str).tolist()

                if pd.api.types.is_numeric_dtype(col):
                    col_type = "число"
                elif pd.api.types.is_datetime64_any_dtype(col):
                    col_type = "дата"
                elif pd.api.types.is_bool_dtype(col):
                    col_type = "логический"
                else:
                    # Check if string values look like dates
                    is_date = False
                    for s in samples[:3]:
                        if '-' in s and len(s) >= 8:
                            try:
                                pd.to_datetime(s)
                                is_date = True
                                break
                            except (ValueError, TypeError):
                                pass
                    col_type = "дата" if is_date else "текст"

                columns.append({
                    "name": str(col_name),
                    "type": col_type,
                    "sample_values": samples[:3],
                })

            # Preview rows (first 10)
            preview_df = df.head(10)
            preview_rows = []
            for _, row in preview_df.iterrows():
                row_dict = {}
                for col_name in df.columns:
                    val = row[col_name]
                    # Convert NaN/NaT to None for JSON
                    if pd.isna(val):
                        row_dict[str(col_name)] = None
                    else:
                        row_dict[str(col_name)] = str(val) if not isinstance(val, (int, float, bool)) else val
                preview_rows.append(row_dict)

            sheets.append(ExcelSheetInfo(
                name=str(sheet_name),
                rows_count=len(df),
                columns=columns,
                preview_rows=preview_rows,
            ))

        return ExcelAnalysisResult(
            sheet_names=sheet_names,
            sheets=sheets,
            total_rows=total_rows,
        )

    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="pandas and openpyxl are required for Excel analysis"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel analysis failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Excel analysis failed: {str(e)}"
        )


# ============================================
# Excel Preview — raw cell data for spreadsheet view
# ============================================

MAX_PREVIEW_ROWS = 100
MAX_PREVIEW_COLS = 30


class SheetPreviewData(BaseModel):
    """Raw cell data for one sheet — for spreadsheet-like rendering."""
    name: str
    max_row: int
    max_col: int
    cells: list[list]  # 2D array of cell values (row-major, 0-indexed)
    visible_rows: int
    visible_cols: int


class ExcelPreviewResponse(BaseModel):
    """Excel file preview with raw cell data for all sheets."""
    sheets: list[SheetPreviewData]


def _safe_cell_value(val):
    """Convert cell value to JSON-safe format."""
    if val is None:
        return None
    import datetime
    if isinstance(val, datetime.datetime):
        return val.isoformat()
    if isinstance(val, (int, float, bool)):
        return val
    return str(val)


@router.post("/{file_id}/excel-preview", response_model=ExcelPreviewResponse)
async def excel_preview(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return raw cell data for spreadsheet-like preview rendering.

    Returns cell values as 2D arrays for each sheet, limited to
    MAX_PREVIEW_ROWS × MAX_PREVIEW_COLS.
    Used by ExcelSourceDialog to render interactive spreadsheet grid
    where users can select table regions.
    """
    try:
        import openpyxl

        storage = get_storage()
        file_content = await storage.get(file_id=file_id, db=db)

        if file_content is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")

        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)

        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            cells_dict: dict[tuple[int, int], any] = {}
            max_row = 0
            max_col = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        r, c = cell.row, cell.column
                        cells_dict[(r, c)] = cell.value
                        max_row = max(max_row, r)
                        max_col = max(max_col, c)

            visible_rows = min(max_row, MAX_PREVIEW_ROWS)
            visible_cols = min(max_col, MAX_PREVIEW_COLS)

            cells_2d: list[list] = []
            for r in range(1, visible_rows + 1):
                row_data = []
                for c in range(1, visible_cols + 1):
                    row_data.append(_safe_cell_value(cells_dict.get((r, c))))
                cells_2d.append(row_data)

            sheets.append(SheetPreviewData(
                name=sheet_name,
                max_row=max_row,
                max_col=max_col,
                cells=cells_2d,
                visible_rows=visible_rows,
                visible_cols=visible_cols,
            ))

        wb.close()
        return ExcelPreviewResponse(sheets=sheets)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel preview failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Excel preview failed: {str(e)}",
        )


# ============================================
# Smart Excel Analysis — table detection
# ============================================

class DetectedRegionResponse(BaseModel):
    """Detected table region within a sheet."""
    sheet_name: str
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    header_row: int | None = None
    confidence: float
    table_name: str
    columns: list[dict]
    preview_rows: list[dict]
    row_count: int
    range_str: str
    detection_method: str


class SheetDetectionResponse(BaseModel):
    """Detection result for one sheet."""
    sheet_name: str
    total_rows: int
    total_cols: int
    regions: list[DetectedRegionResponse]
    grid_map: list[list[str]]
    grid_rows: int
    grid_cols: int


class SmartExcelAnalysisResult(BaseModel):
    """Smart Excel analysis result with detected table regions."""
    sheet_names: list[str]
    sheets: list[SheetDetectionResponse]
    total_tables_found: int
    detection_method: str


@router.post("/{file_id}/analyze-excel-smart", response_model=SmartExcelAnalysisResult)
async def analyze_excel_smart(
    file_id: str,
    use_ai: bool = True,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Smart Excel analysis — detect tables in arbitrary positions.
    
    Uses hybrid approach: heuristic detection + optional AI refinement via GigaChat.
    Returns detected regions (tables) with coordinates, columns, preview data,
    and a grid_map for visual rendering on frontend.
    
    Args:
        file_id: Uploaded file ID
        use_ai: Whether to use AI (GigaChat) for refinement (default: True)
    
    See docs/SOURCE_NODE_CONCEPT_V2.md — Excel Dialog.
    """
    try:
        storage = get_storage()
        file_content = await storage.get(file_id=file_id, db=db)

        if file_content is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found",
            )

        # Get GigaChat service if AI requested
        gigachat_service = None
        if use_ai:
            try:
                from app.services.gigachat_service import get_gigachat_service
                gigachat_service = get_gigachat_service()
            except Exception as e:
                logger.warning(f"GigaChat not available for smart analysis, falling back to heuristic: {e}")

        # Run detection
        from app.sources.excel.table_detector import ExcelTableDetector
        detector = ExcelTableDetector(gigachat_service=gigachat_service)
        result = await detector.detect(file_content)

        return SmartExcelAnalysisResult(**result.to_dict())

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Smart Excel analysis failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Smart Excel analysis failed: {str(e)}",
        )


# ============================================
# Document Analysis
# ============================================

class DocumentAnalysisResult(BaseModel):
    """Document analysis result — text + tables extracted from PDF/DOCX/TXT."""
    document_type: str  # pdf, docx, txt
    filename: str
    text: str  # Extracted text (or truncated preview)
    text_length: int  # Full text length
    page_count: int | None = None  # For PDFs
    tables: list[dict]  # [{name, columns: [{name, type}], rows: [{col: val}], row_count}]
    table_count: int
    total_rows: int  # Sum of rows across all tables
    is_scanned: bool = False  # True if PDF has no text layer


@router.post("/{file_id}/analyze-document", response_model=DocumentAnalysisResult)
async def analyze_document_file(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Analyze a document file (PDF/DOCX/TXT) and return extracted text + tables.
    
    Similar to analyze-csv but for documents. Returns:
    - Extracted text (preview up to 5000 chars)
    - Tables found in the document
    - Document metadata (page count, type, etc.)
    
    See docs/SOURCE_NODE_CONCEPT_V2.md — раздел "📄 4. Document Dialog".
    """
    try:
        storage = get_storage()
        file_content = await storage.get(file_id=file_id, db=db)
        
        if file_content is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found"
            )
        
        # Resolve filename from DB
        from app.models import UploadedFile
        from sqlalchemy import select
        from uuid import UUID as PyUUID
        
        result = await db.execute(
            select(UploadedFile).where(UploadedFile.id == PyUUID(file_id))
        )
        uploaded_file = result.scalar_one_or_none()
        filename = uploaded_file.filename if uploaded_file else "document"
        mime_type = uploaded_file.mime_type if uploaded_file else ""
        
        # Run document extraction
        from app.sources.document.extractor import DocumentSource, detect_document_type
        
        config = {
            "file_id": file_id,
            "filename": filename,
            "mime_type": mime_type,
        }
        doc_type = detect_document_type(config)
        config["document_type"] = doc_type
        
        extractor = DocumentSource()
        extraction = await extractor.extract(config, file_content=file_content)
        
        if not extraction.success:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=extraction.error or "Document extraction failed"
            )
        
        # Prepare tables for response
        tables_data = []
        for t in extraction.tables:
            preview_rows = t.rows[:20]  # First 20 rows as preview
            tables_data.append({
                "name": t.name,
                "columns": t.columns,
                "rows": preview_rows,
                "row_count": len(t.rows),
            })
        
        total_rows = sum(len(t.rows) for t in extraction.tables)
        is_scanned = "не содержит текстового слоя" in extraction.text if extraction.text else False
        
        # Return text preview (up to 5000 chars)
        text_preview = extraction.text[:5000] if extraction.text else ""
        
        return DocumentAnalysisResult(
            document_type=doc_type,
            filename=filename,
            text=text_preview,
            text_length=len(extraction.text) if extraction.text else 0,
            page_count=extraction.metadata.get("page_count"),
            tables=tables_data,
            table_count=len(extraction.tables),
            total_rows=total_rows,
            is_scanned=is_scanned,
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Document analysis failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Document analysis failed: {str(e)}"
        )


# ============================================
# Document Extraction Chat (Multi-Agent)
# ============================================

class DocumentExtractionChatRequest(BaseModel):
    """Request for iterative document extraction via multi-agent chat."""
    user_prompt: str
    document_text: str  # Full extracted text from analyze-document
    document_type: str = "txt"  # pdf | docx | txt
    filename: str = "document"
    page_count: int | None = None
    existing_tables: list[dict] = []  # Already extracted tables from previous turns
    chat_history: list[dict] = []  # [{role: "user"|"assistant", content: str}]


class DocumentExtractionChatResponse(BaseModel):
    """Response from document extraction chat."""
    narrative: str  # AI response text (markdown)
    tables: list[dict] = []  # Extracted/updated tables
    findings: list[dict] = []  # Extracted findings
    status: str = "success"
    mode: str = "document_extraction"
    agent_plan: dict | None = None


@router.post("/{file_id}/extract-document-chat", response_model=DocumentExtractionChatResponse)
async def extract_document_chat(
    file_id: str,
    request: DocumentExtractionChatRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Iterative AI chat for extracting structured data from documents.
    
    Uses DocumentExtractionController → Orchestrator V2.
    Each message refines extraction — tables update in real-time.
    
    See docs/SOURCE_NODE_CONCEPT_V2.md — "📄 4. Document Dialog"
    """
    if not request.user_prompt:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="user_prompt is required"
        )

    try:
        # Get Orchestrator
        from app.main import get_orchestrator
        orchestrator = get_orchestrator()
        if not orchestrator:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Orchestrator not initialized. Check backend logs for Redis/GigaChat connection issues."
            )

        from app.services.controllers.document_extraction_controller import DocumentExtractionController
        controller = DocumentExtractionController(orchestrator)

        result = await controller.process_request(
            user_message=request.user_prompt,
            context={
                "board_id": "",  # Document extraction is file-level, not board-level
                "user_id": str(current_user.id),
                "document_text": request.document_text,
                "document_type": request.document_type,
                "filename": request.filename,
                "existing_tables": request.existing_tables,
                "chat_history": request.chat_history,
                "page_count": request.page_count,
            },
        )

        if result.status != "success":
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=result.error or "Document extraction failed"
            )

        return DocumentExtractionChatResponse(
            narrative=result.narrative or "Анализ выполнен.",
            tables=result.metadata.get("tables", []),
            findings=result.metadata.get("findings", []),
            status="success",
            mode=result.mode or "document_extraction",
            agent_plan=result.plan,
        )

    except HTTPException:
        raise
    except RuntimeError as e:
        error_msg = str(e)
        if "Failed to get response from GigaChat" in error_msg or "getaddrinfo failed" in error_msg:
            logger.error(f"GigaChat API connection error: {e}")
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Не удалось подключиться к GigaChat API. Проверьте GIGACHAT_CREDENTIALS в .env"
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Document extraction failed: {str(e)}"
        )
    except Exception as e:
        logger.error(f"Document extraction chat failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Document extraction failed: {str(e)}"
        )
